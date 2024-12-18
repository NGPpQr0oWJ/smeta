from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, jsonify
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import pandas as pd
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)



app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
MATERIALS_PATH = os.path.join('database', 'materials.db')

DB_PATHS = {
    "main": "tech_map.db",
    "materials": os.path.join('database', 'materials.db')
}

def get_db_connection(db_path):
    """Создает подключение к базе данных SQLite."""
    return sqlite3.connect(db_path)

def init_tech_map_db():
    """Инициализация базы данных tech_map.db."""
    conn = get_db_connection(DB_PATHS["main"])
    cursor = conn.cursor()

    # Создание таблицы пользователей
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        );
    ''')

    # Создание таблицы продуктов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            order_number TEXT NOT NULL,
            creation_date TEXT NOT NULL,
            project_file TEXT
        );
    ''')

    conn.commit()
    conn.close()

def init_materials_db():
    """Инициализация базы данных materials.db."""
    conn = get_db_connection(DB_PATHS["materials"])
    cursor = conn.cursor()

    # Создание таблицы материалов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            article TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            unit TEXT NOT NULL
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            article TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            unit TEXT NOT NULL
        );
    ''')

    conn.commit()
    conn.close()


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Страница авторизации пользователя."""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('tech_map.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
        user = cursor.fetchone()
        conn.close()
        if user:
            session['user_id'] = user[0]
            session['username'] = user[1]
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Неверное имя пользователя или пароль')
    return render_template('login.html')
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@app.route('/update_materials', methods=['POST'])
def update_materials():
    """Добавление нового материала с проверкой уникальности артикула."""
    data = request.json
    article = data.get('Артикул')
    material = data.get('Материал')
    unit = data.get('Ед. изм.')

    if not all([article, material, unit]):
        return jsonify({"status": "error", "message": "Заполните все поля"}), 400

    try:
        conn = get_db_connection(DB_PATHS["materials"])
        cursor = conn.cursor()

        # Проверка уникальности
        cursor.execute("SELECT name FROM materials WHERE article = ?", (article,))
        if cursor.fetchone():
            return jsonify({"status": "error", "message": f"Артикул {article} уже существует"}), 400

        # Добавление материала
        cursor.execute("INSERT INTO materials (article, name, unit) VALUES (?, ?, ?)", (article, material, unit))
        conn.commit()
    except Exception as e:
        return jsonify({"status": "error", "message": f"Ошибка: {str(e)}"}), 500
    finally:
        conn.close()

    return jsonify({"status": "success", "message": "Материал успешно добавлен"}), 200

@app.route('/logout')
def logout():
    """Выход пользователя."""
    session.pop('user_id', None)
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    """Добавление нового пользователя администратором."""
    if 'user_id' not in session or session.get('username') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        fio = request.form['fio']
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('tech_map.db')
        cursor = conn.cursor()
        try:
            cursor.execute('INSERT INTO users (username, password) VALUES (?, ?)', (username, password))
            conn.commit()
            conn.close()
            return redirect(url_for('index'))
        except sqlite3.IntegrityError:
            conn.close()
            return render_template('add_user.html', error='Имя пользователя уже занято')
    return render_template('add_user.html')

@app.route('/')
def index():
    """Главная страница: список всех продуктов."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    is_admin = session.get('username') == 'admin'
    search_query = request.args.get('search', '')
    sort_by_date = request.args.get('sort_by_date', '')

    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()

    if search_query:
        cursor.execute('''
            SELECT * FROM products
            WHERE name LIKE ? OR order_number LIKE ?
            ORDER BY creation_date DESC
        ''', (f'%{search_query}%', f'%{search_query}%'))
    else:
        cursor.execute('SELECT * FROM products ORDER BY creation_date DESC' if sort_by_date else 'SELECT * FROM products')

    products = cursor.fetchall()
    conn.close()

    return render_template('index.html', products=products, is_admin=is_admin, search_query=search_query, sort_by_date=sort_by_date)

@app.route('/product/<int:product_id>')
def product_details(product_id):
    """Детали продукта: список материалов."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    cursor.execute('SELECT * FROM materials WHERE product_id = ?', (product_id,))
    materials = cursor.fetchall()
    conn.close()
    return render_template('product_details.html', product=product, materials=materials)

@app.route('/delete_product/<int:product_id>', methods=['POST'])
def delete_product(product_id):
    """Удаление продукта."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM products WHERE id = ?', (product_id,))
    cursor.execute('DELETE FROM materials WHERE product_id = ?', (product_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/add_product', methods=['GET', 'POST'])
def add_product():
    """Добавление нового продукта."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        name = request.form['name']
        order_number = request.form['order_number']
        creation_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        project_file = None
        if 'project_file' in request.files:
            file = request.files['project_file']
            if file and file.filename != '' and file.content_length <= 100 * 1024 * 1024:
                project_file = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(project_file)
        conn = sqlite3.connect('tech_map.db')
        cursor = conn.cursor()
        cursor.execute('INSERT INTO products (name, order_number, creation_date, project_file) VALUES (?, ?, ?, ?)', (name, order_number, creation_date, project_file))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    return render_template('add_product.html', datetime=datetime)

@app.route('/download_project/<int:product_id>')
def download_project(product_id):
    """Скачивание проектной документации."""
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute('SELECT project_file FROM products WHERE id = ?', (product_id,))
    project_file = cursor.fetchone()[0]
    conn.close()
    if project_file and os.path.exists(project_file):
        return send_from_directory(app.config['UPLOAD_FOLDER'], os.path.basename(project_file), as_attachment=True)
    return redirect(request.referrer)

@app.route('/add_material/<int:product_id>', methods=['POST'])
def add_material(product_id):
    """Добавление материала к продукту."""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    material = request.form.get('material')
    article = request.form.get('article')
    unit = request.form.get('unit')
    quantity = request.form.get('quantity')
    notes = request.form.get('notes')

    # Проверяем, что обязательные поля заполнены
    if not material or not quantity:
        return redirect(url_for('product_details', product_id=product_id))

    # Добавляем материал в базу данных
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute(
        'INSERT INTO materials (product_id, name, article, quantity, unit, notes) VALUES (?, ?, ?, ?, ?, ?)',
        (product_id, material, article, quantity, unit, notes)
    )
    conn.commit()
    conn.close()
    return redirect(url_for('product_details', product_id=product_id))

@app.route('/delete_material/<int:material_id>', methods=['POST'])
def delete_material(material_id):
    """Удаление материала."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM materials WHERE id = ?', (material_id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer)

import sqlite3
import pandas as pd
import os
from flask import request, jsonify

MATERIALS_PATH = os.path.join('database', 'materials.db')

@app.route('/import_materials', methods=['POST'])
def import_materials():
    """Импорт материалов из Excel с проверкой уникальности."""
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Файл не загружен"}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "Неверный формат файла. Требуется .xlsx"}), 400

    import pandas as pd
    conn = sqlite3.connect(MATERIALS_PATH)
    cursor = conn.cursor()
    inserted, skipped = 0, 0

    try:
        df = pd.read_excel(file)
        for _, row in df.iterrows():
            article = row['Артикул']
            material = row['Материал']
            unit = row['Ед. изм.']

            # Проверка уникальности артикула
            cursor.execute("SELECT COUNT(*) FROM materials WHERE article = ?", (article,))
            if cursor.fetchone()[0] == 0:
                cursor.execute(
                    "INSERT INTO materials (article, name, unit) VALUES (?, ?, ?)",
                    (article, material, unit)
                )
                inserted += 1
            else:
                skipped += 1

        conn.commit()
    except Exception as e:
        return jsonify({"status": "error", "message": f"Ошибка: {str(e)}"}), 500
    finally:
        conn.close()

    return jsonify({"status": "success", "message": f"Импорт завершён. Добавлено: {inserted}, пропущено: {skipped}."})

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

    
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

@app.route('/export_estimate/<int:product_id>', methods=['GET'])
def export_estimate(product_id):
    """Экспорт сметы изделия в Excel-файл."""
    # Подключение к базе данных и получение данных о продукте и материалах
    conn = sqlite3.connect('tech_map.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
    product = cursor.fetchone()

    cursor.execute('SELECT * FROM materials WHERE product_id = ?', (product_id,))
    materials = cursor.fetchall()
    conn.close()

    # Создание рабочей книги
    wb = openpyxl.Workbook()
    ws = wb.active

    # Название листа
    ws.title = "Смета"

    # Формирование названия файла
    current_date = datetime.now().strftime('%Y-%m-%d')
    file_name = f"{product[1]}_{product[2]}_{current_date}.xlsx"

    # Шапка отчета
    ws['A1'] = "Расчет стоимости материалов и комплектующих на "
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['C1'] = current_date  # Текущая дата
    ws['C1'].font = Font(bold=True, color="FF0000")

    ws['A2'] = "Таблица Общая"
    ws.merge_cells('A2:G2')
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Данные изделия
    ws['A4'] = "Заказ"
    ws['B4'] = product[2]  # Номер заказа
    ws['B4'].font = Font(bold=True, color="FF0000")

    ws['A5'] = "Изделие"
    ws['B5'] = product[1]  # Название изделия
    ws['B5'].font = Font(bold=True, color="FF0000")

    ws['A8'] = "Разработчик"
    ws['B8'] = session.get('username', "Неизвестный пользователь")  # Текущий пользователь
    ws['B8'].font = Font(bold=True, color="FF0000")

    # Скрытие колонок E и F
    ws.column_dimensions['E'].hidden = True
    ws.column_dimensions['F'].hidden = True

    # Шапка таблицы материалов
    headers = ["Артикул", "Наименование материала", "Ед. изм.", "Количество", "", "", "Примечание"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Стиль рамки
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем рамки к шапке материалов (A11:G11)
    for col in range(1, 8):
        ws.cell(row=11, column=col).border = thin_border

    # Заполнение данных материалов с правильным позиционированием
    row_start = 12
    for row_num, material in enumerate(materials, row_start):
        ws.cell(row=row_num, column=1, value=material[2])  # Артикул
        ws.cell(row=row_num, column=2, value=material[3])  # Материал
        ws.cell(row=row_num, column=3, value=material[5])  # Количество
        ws.cell(row=row_num, column=4, value=material[4])  # Ед. изм.
        ws.cell(row=row_num, column=7, value=material[6])  # Примечание

        # Добавляем рамки к строкам материалов
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = thin_border

    # Автоматическое выравнивание по ширине колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву колонки
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохранение файла
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
    wb.save(save_path)

    # Отправка файла пользователю
    return send_from_directory(app.config['UPLOAD_FOLDER'], file_name, as_attachment=True)

@app.route("/")
def home():
    return "Hello, Flask!"

if __name__ == '__main__':
    init_tech_map_db()
    init_materials_db()
    app.run(debug=True)

